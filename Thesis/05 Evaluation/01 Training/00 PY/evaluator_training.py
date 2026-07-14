import json, re, sys
from dataclasses import dataclass, field
from pathlib import Path

def find_project_root(marker: str = ".git") -> Path:
    """Laeuft vom Skriptpfad aus nach oben, bis der Marker (Standard: .git-Ordner)
    gefunden wird. Funktioniert unabhaengig von Nutzername, Laufwerk und davon,
    wie tief das Skript im Ordnerbaum liegt."""
    p = Path(__file__).resolve()
    for parent in p.parents:
        if (parent / marker).exists():
            return parent
    raise RuntimeError(f"Projekt-Root nicht gefunden (kein '{marker}' oberhalb von {p})")


PROJECT_ROOT = find_project_root()

from typing import Any, get_args, get_origin

import numpy as np
from openpyxl import Workbook
from scipy.optimize import linear_sum_assignment

sys.path.append(str(PROJECT_ROOT / "01 Schema"))
from epd import EPD

ROOT                = Path(PROJECT_ROOT)
GROUND_TRUTH_PATH   = ROOT / "05 Evaluation" / "00 Ground Truth" / "ground_truth_training.json"
EXTRACTION_ROOT     = ROOT / "03 Extraktion" / "01 Training"
OUTPUT_ROOT         = ROOT / "05 Evaluation" / "01 Training"
VARIANTS = ["01 pdfplumber plain", "03 pymupdf4llm plain", "05 pdfplumber prompt optimized", "06 pymupdf4llm prompt optimized"]

def _infer_type(a) -> str:
    inner = (get_args(a) or (a,))[0]
    if get_origin(inner) is list: return "list"
    if inner is bool:             return "bool"
    if inner in (float, int):     return "numeric"
    return "text"

FIELD_TYPES = {n: _infer_type(f.annotation) for n, f in EPD.model_fields.items()}
ALL_FIELDS  = list(FIELD_TYPES)

def is_empty(v):  return v is None or (isinstance(v, str) and not v.strip()) or (isinstance(v, (list, tuple)) and not v)

_NUM_RE = re.compile(r"[-+]?[\d.,]*\d(?:[eE][-+]?\d+)?")

def to_float(v):
    if isinstance(v, (int, float)):
        return float(v)
    m = _NUM_RE.search(str(v))
    if m is None:
        raise ValueError(f"keine Zahl gefunden in {v!r}")
    return float(m.group().rstrip(".,").replace(",", "."))

def are(g, p):    g, p = to_float(g), to_float(p); m = max(abs(g), abs(p)); return 0.0 if m == 0 else abs(g - p) / m

def levenshtein(a, b):
    if a == b: return 0
    if not a:  return len(b)
    if not b:  return len(a)
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        curr = [i] + [0] * len(b)
        for j, cb in enumerate(b, 1):
            curr[j] = min(prev[j-1] + (ca != cb), prev[j] + 1, curr[j-1] + 1)
        prev = curr
    return prev[-1]

def nld(a, b):
    a, b = str(a), str(b)
    m = max(len(a), len(b))
    return 0.0 if m == 0 else levenshtein(a, b) / m

def list_dist(gt, pred):
    gt, pred = [str(x) for x in gt], [str(x) for x in pred]
    n, m = len(gt), len(pred)
    if n == 0 and m == 0: return 0.0
    if n == 0 or m == 0:  return 1.0
    cost = np.array([[nld(g, p) for p in pred] for g in gt])
    ri, ci = linear_sum_assignment(cost)
    s = float(sum((1 - cost)[i, j] for i, j in zip(ri, ci)))
    l = float(len(ri) + (m - len(set(ci.tolist()))) + (n - len(set(ri.tolist()))))
    return 0.0 if l == 0 else 1.0 - s / l

@dataclass
class FieldResult:
    document_id: str; field: str; type: str
    ground_truth: Any; extracted: Any; status: str
    are: float | None = None; nld: float | None = None; bool_match: int | None = None

@dataclass
class FieldSummary:
    field: str; type: str
    mean_are: float | None = None; median_are: float | None = None
    mean_nld: float | None = None; median_nld: float | None = None
    bool_accuracy: float | None = None
    n_present: int = 0; n_omission: int = 0; n_hallucination: int = 0; n_absent: int = 0

@dataclass
class VariantenMetriken:
    mean_are: float | None = None; median_are: float | None = None
    mean_nld: float | None = None; median_nld: float | None = None


def load_ground_truth(path: Path) -> dict:
    data = json.loads(path.read_text(encoding="utf-8"))
    epds = data["epd"] if isinstance(data, dict) and "epd" in data else data
    return {e["document_id"]: e for e in epds}

def load_extractions(folder: Path) -> dict:
    return {p.stem: json.loads(p.read_text(encoding="utf-8")) for p in sorted(folder.rglob("*.json"))}

def find_variant_folder(variant: str) -> Path:
    hits = list(EXTRACTION_ROOT.glob(f"*{variant}"))
    if not hits: raise FileNotFoundError(f"Kein Ordner fuer '{variant}' unter {EXTRACTION_ROOT}.")
    return hits[0]

def evaluate_field(doc_id, field, ftype, gt, pred) -> FieldResult:
    ge, pe = is_empty(gt), is_empty(pred)
    status = "absent" if ge and pe else "hallucination" if ge else "omission" if pe else "present"
    r = FieldResult(doc_id, field, ftype, gt, pred, status)
    if status != "present":
        if ftype == "bool": r.bool_match = 0
        return r
    if   ftype == "numeric":
        try: r.are = are(gt, pred)
        except (TypeError, ValueError) as e: print(f"  [WARNUNG] {doc_id}/{field}: {e}")
    elif ftype == "text":    r.nld = nld(gt, pred)
    elif ftype == "bool":    r.bool_match = 1 if bool(gt) == bool(pred) else 0
    elif ftype == "list":    r.nld = list_dist(gt if isinstance(gt, list) else [gt], pred if isinstance(pred, list) else [pred])
    return r

def evaluate_variant(variant, gt_all) -> list[FieldResult]:
    extractions = load_extractions(find_variant_folder(variant))
    results, missing = [], []
    for doc_id, gt_epd in gt_all.items():
        pred = extractions.get(doc_id)
        if pred is None: missing.append(doc_id); continue
        for f in ALL_FIELDS:
            results.append(evaluate_field(doc_id, f, FIELD_TYPES[f], gt_epd.get(f), pred.get(f)))
    if missing: print(f"  [WARNUNG] '{variant}': keine Datei fuer {missing}.")
    return results

def summarize(results: list[FieldResult]) -> tuple[list[FieldSummary], VariantenMetriken]:
    sums   = {f: FieldSummary(field=f, type=FIELD_TYPES[f]) for f in ALL_FIELDS}
    are_v  = {f: [] for f in ALL_FIELDS}
    nld_v  = {f: [] for f in ALL_FIELDS}
    bool_v = {f: [] for f in ALL_FIELDS}
    for r in results:
        s = sums[r.field]
        match r.status:
            case "present":
                s.n_present += 1
                if r.are        is not None: are_v[r.field].append(r.are)
                if r.nld        is not None: nld_v[r.field].append(r.nld)
                if r.bool_match is not None: bool_v[r.field].append(r.bool_match)
            case "omission":
                s.n_omission += 1
                if r.bool_match is not None: bool_v[r.field].append(r.bool_match)
            case "hallucination":
                s.n_hallucination += 1
                if r.bool_match is not None: bool_v[r.field].append(r.bool_match)
            case "absent":
                s.n_absent += 1
    for f, s in sums.items():
        if are_v[f]:  s.mean_are = float(np.mean(are_v[f])); s.median_are = float(np.median(are_v[f]))
        if nld_v[f]:  s.mean_nld = float(np.mean(nld_v[f])); s.median_nld = float(np.median(nld_v[f]))
        if bool_v[f]: s.bool_accuracy = float(np.mean(bool_v[f]))
    alle_are = [v for f in ALL_FIELDS for v in are_v[f]]
    alle_nld = [v for f in ALL_FIELDS for v in nld_v[f]]
    return list(sums.values()), VariantenMetriken(
        mean_are=float(np.mean(alle_are)) if alle_are else None,
        median_are=float(np.median(alle_are)) if alle_are else None,
        mean_nld=float(np.mean(alle_nld)) if alle_nld else None,
        median_nld=float(np.median(alle_nld)) if alle_nld else None,
    )

def _r(v, d=6): return round(v, d) if v is not None else ""
def _fmt(v):    return "" if v is None else (json.dumps(v, ensure_ascii=False) if isinstance(v, (list, dict)) else v)

def write_variant_excel(results, summaries, path: Path) -> None:
    wb = Workbook()
    ws1 = wb.active; ws1.title = "Details"
    ws1.append(["document_id","field","type","ground_truth","extracted","status","are","nld","bool_match"])
    for r in results:
        ws1.append([r.document_id, r.field, r.type, _fmt(r.ground_truth), _fmt(r.extracted),
                    r.status, _r(r.are), _r(r.nld), r.bool_match if r.bool_match is not None else ""])
    ws2 = wb.create_sheet("Summary")
    ws2.append(["field","type","mean_are","median_are","mean_nld","median_nld","bool_accuracy","n_present","n_omission","n_hallucination","n_absent"])
    for s in summaries:
        ws2.append([s.field, s.type, _r(s.mean_are), _r(s.median_are), _r(s.mean_nld), _r(s.median_nld), _r(s.bool_accuracy),
                    s.n_present, s.n_omission, s.n_hallucination, s.n_absent])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path); print(f"  -> {path}")

def write_comparison_excel(all_sums, all_vm, path: Path) -> None:
    wb  = Workbook()
    metrics = ["mean_are","median_are","mean_nld","median_nld","bool_accuracy","n_present","n_omission","n_hallucination","n_absent"]
    ws1 = wb.active; ws1.title = "Vergleich"
    ws1.append(["field","type"] + [f"{v}_{m}" for v in VARIANTS for m in metrics])
    sd  = {v: {s.field: s for s in sums} for v, sums in all_sums.items()}
    for f in ALL_FIELDS:
        row = [f, FIELD_TYPES[f]]
        for v in VARIANTS:
            s = sd[v].get(f)
            row += [""] * len(metrics) if s is None else [
                _r(s.mean_are), _r(s.median_are), _r(s.mean_nld), _r(s.median_nld), _r(s.bool_accuracy),
                s.n_present, s.n_omission, s.n_hallucination, s.n_absent]
        ws1.append(row)
    ws2 = wb.create_sheet("Gesamt")
    ws2.append(["variant","mean_are","median_are","mean_nld","median_nld","bool_accuracy_all",
                "total_present","total_omission","total_hallucination","total_absent"])
    avg = lambda lst: float(np.mean(lst)) if lst else None
    for v, sums in all_sums.items():
        vm = all_vm[v]
        ws2.append([v, _r(vm.mean_are), _r(vm.median_are), _r(vm.mean_nld), _r(vm.median_nld),
            _r(avg([s.bool_accuracy for s in sums if s.bool_accuracy is not None])),
            sum(s.n_present for s in sums), sum(s.n_omission for s in sums),
            sum(s.n_hallucination for s in sums), sum(s.n_absent for s in sums)])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path); print(f"  -> {path}")

def main() -> None:
    print("Lade Ground Truth (Training) ...")
    gt = load_ground_truth(GROUND_TRUTH_PATH)
    print(f"  {len(gt)} EPDs.\n")
    all_sums, all_vm = {}, {}
    for variant in VARIANTS:
        print(f"Evaluiere: {variant}")
        try:
            results = evaluate_variant(variant, gt)
        except FileNotFoundError as e:
            print(f"  [FEHLER] {e}\n"); continue
        sums, vm = summarize(results)
        all_sums[variant], all_vm[variant] = sums, vm
        write_variant_excel(results, sums, OUTPUT_ROOT / variant / f"evaluation_{variant}.xlsx")
        print()
    if len(all_sums) > 1:
        print("Schreibe Vergleichsdatei ...")
        write_comparison_excel(all_sums, all_vm, OUTPUT_ROOT / "evaluation_comparison.xlsx")
    print("\nFertig.")

if __name__ == "__main__":
    main()
